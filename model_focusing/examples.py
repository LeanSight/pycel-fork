#!/usr/bin/env python3
# -*- coding: UTF-8 -*-
"""
Ejemplos prácticos de uso de las capacidades de Model Focusing en Pycel.

Este archivo contiene ejemplos reales de cómo utilizar las funcionalidades core
de model focusing para análisis industrial de planillas Excel.
"""

import os
import tempfile
from pathlib import Path

from pycel import ExcelCompiler


def example_1_financial_model_audit():
    """
    Ejemplo 1: Auditoría de Modelo Financiero
    
    Demuestra cómo extraer y validar un sub-modelo crítico
    de un modelo financiero complejo.
    """
    print("=== Ejemplo 1: Auditoría de Modelo Financiero ===")
    
    # Simular carga de modelo financiero
    # En la práctica, sería: ExcelCompiler('financial_model.xlsx')
    excel = create_sample_financial_model()
    
    print(f"Modelo original: {len(excel.cell_map)} celdas")
    
    # Definir inputs y outputs críticos
    input_addrs = [
        'Assumptions!B1',  # Revenue_Growth
        'Assumptions!B2',  # COGS_Rate
        'Assumptions!B3'   # OpEx_Rate
    ]
    
    output_addrs = [
        'Summary!B1',  # Revenue
        'Summary!B5',  # EBITDA
        'Summary!B6'   # FCF
    ]
    
    print(f"Inputs críticos: {input_addrs}")
    print(f"Outputs críticos: {output_addrs}")
    
    # Extraer sub-modelo
    excel.trim_graph(input_addrs=input_addrs, output_addrs=output_addrs)
    
    print(f"Modelo trimmed: {len(excel.cell_map)} celdas")
    print(f"Reducción: {((len(excel.cell_map) / len(excel.cell_map)) * 100):.1f}%")
    
    # Validar precisión del sub-modelo
    validation_results = excel.validate_calcs(output_addrs=output_addrs)
    
    if validation_results:
        print("⚠️ Discrepancias encontradas:")
        for category, errors in validation_results.items():
            print(f"  {category}: {len(errors)} errores")
    else:
        print("✅ Sub-modelo validado correctamente")
    
    # Mostrar valores actuales
    print("\nValores actuales:")
    for addr in output_addrs:
        try:
            value = excel.evaluate(addr)
            print(f"  {addr}: {value:,.2f}" if isinstance(value, (int, float)) else f"  {addr}: {value}")
        except:
            print(f"  {addr}: No disponible")
    
    return excel


def example_2_sensitivity_analysis():
    """
    Ejemplo 2: Análisis de Sensibilidad
    
    Demuestra cómo realizar análisis de sensibilidad
    modificando inputs y observando impacto en outputs.
    """
    print("\n=== Ejemplo 2: Análisis de Sensibilidad ===")
    
    # Usar modelo del ejemplo anterior
    excel = create_sample_financial_model()
    
    # Extraer sub-modelo para análisis
    input_addrs = ['Assumptions!B1', 'Assumptions!B2']  # Revenue_Growth, COGS_Rate
    output_addrs = ['Summary!B5']  # EBITDA
    
    excel.trim_graph(input_addrs=input_addrs, output_addrs=output_addrs)
    
    # Definir escenarios de sensibilidad
    scenarios = [
        {'name': 'Base Case', 'Revenue_Growth': 0.05, 'COGS_Rate': 0.60},
        {'name': 'Optimistic', 'Revenue_Growth': 0.10, 'COGS_Rate': 0.55},
        {'name': 'Pessimistic', 'Revenue_Growth': 0.02, 'COGS_Rate': 0.65},
        {'name': 'High Growth', 'Revenue_Growth': 0.15, 'COGS_Rate': 0.60},
        {'name': 'Cost Pressure', 'Revenue_Growth': 0.05, 'COGS_Rate': 0.70},
    ]
    
    print("Análisis de Escenarios:")
    print("Scenario".ljust(15) + "Rev Growth".ljust(12) + "COGS Rate".ljust(12) + "EBITDA")
    print("-" * 55)
    
    results = []
    for scenario in scenarios:
        # Configurar inputs del escenario
        excel.set_value('Assumptions!B1', scenario['Revenue_Growth'])
        excel.set_value('Assumptions!B2', scenario['COGS_Rate'])
        
        # Evaluar output
        try:
            ebitda = excel.evaluate('Summary!B5')
            results.append({
                'scenario': scenario['name'],
                'ebitda': ebitda,
                'revenue_growth': scenario['Revenue_Growth'],
                'cogs_rate': scenario['COGS_Rate']
            })
            
            print(f"{scenario['name']:<15}{scenario['Revenue_Growth']:<12.1%}{scenario['COGS_Rate']:<12.1%}{ebitda:>10,.0f}")
        except Exception as e:
            print(f"{scenario['name']:<15}Error: {str(e)}")
    
    # Análisis de sensibilidad individual
    print("\nAnálisis de Sensibilidad Individual (Revenue Growth):")
    excel.set_value('Assumptions!B2', 0.60)  # Fijar COGS
    
    growth_rates = [0.00, 0.02, 0.05, 0.08, 0.10, 0.12, 0.15]
    print("Growth Rate".ljust(15) + "EBITDA".ljust(15) + "Change")
    print("-" * 40)
    
    base_ebitda = None
    for rate in growth_rates:
        excel.set_value('Assumptions!B1', rate)
        try:
            ebitda = excel.evaluate('Summary!B5')
            if base_ebitda is None:
                base_ebitda = ebitda
                change = 0
            else:
                change = ebitda - base_ebitda
            
            print(f"{rate:<15.1%}{ebitda:<15,.0f}{change:>+10,.0f}")
        except:
            print(f"{rate:<15.1%}Error")
    
    return results


def example_3_dependency_analysis():
    """
    Ejemplo 3: Análisis de Dependencias
    
    Demuestra cómo analizar y visualizar las dependencias
    de celdas críticas en el modelo.
    """
    print("\n=== Ejemplo 3: Análisis de Dependencias ===")
    
    excel = create_sample_financial_model()
    
    # Analizar dependencias de una celda crítica
    critical_cell = 'Summary!B5'  # EBITDA
    
    print(f"Analizando dependencias de: {critical_cell}")
    
    # Generar árbol de dependencias
    print("\nÁrbol de Dependencias:")
    try:
        for line in excel.value_tree_str(critical_cell):
            print(line)
    except Exception as e:
        print(f"Error generando árbol: {e}")
    
    # Análisis bidireccional
    print(f"\nAnálisis Bidireccional para {critical_cell}:")
    
    try:
        # Evaluar para construir grafo
        excel.evaluate(critical_cell)
        
        target_cell = excel.cell_map.get(critical_cell)
        if target_cell:
            # Precedentes (inputs)
            predecessors = list(excel.dep_graph.predecessors(target_cell))
            print(f"Precedentes directos ({len(predecessors)}):")
            for pred in predecessors[:5]:  # Mostrar solo los primeros 5
                print(f"  ← {pred.address.address}")
            
            # Dependientes (outputs)
            successors = list(excel.dep_graph.successors(target_cell))
            print(f"Dependientes directos ({len(successors)}):")
            for succ in successors[:5]:  # Mostrar solo los primeros 5
                print(f"  → {succ.address.address}")
        else:
            print("Celda no encontrada en el grafo")
            
    except Exception as e:
        print(f"Error en análisis bidireccional: {e}")
    
    # Exportar grafo para análisis visual
    print("\nExportando grafo para análisis visual...")
    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            gexf_file = os.path.join(tmpdir, 'dependency_graph.gexf')
            excel.export_to_gexf(gexf_file)
            print(f"Grafo exportado a: {gexf_file}")
            print("(Puede abrirse en Gephi para visualización)")
    except Exception as e:
        print(f"Error exportando grafo: {e}")


def example_4_model_validation():
    """
    Ejemplo 4: Validación Robusta de Modelo
    
    Demuestra diferentes tipos de validación disponibles
    para asegurar la integridad del modelo.
    """
    print("\n=== Ejemplo 4: Validación Robusta de Modelo ===")
    
    excel = create_sample_financial_model()
    
    # Validación completa del modelo
    print("1. Validación completa del modelo...")
    validation_results = excel.validate_calcs()
    
    print(f"Resultados de validación:")
    if not validation_results:
        print("  ✅ Todos los cálculos son consistentes con Excel")
    else:
        for category, errors in validation_results.items():
            print(f"  ⚠️ {category}: {len(errors)} problemas")
            if isinstance(errors, dict):
                for error_type, error_list in errors.items():
                    print(f"    - {error_type}: {len(error_list)} casos")
    
    # Validación de celdas específicas
    print("\n2. Validación de celdas críticas...")
    critical_cells = ['Summary!B1', 'Summary!B5', 'Summary!B6']  # Revenue, EBITDA, FCF
    
    for cell in critical_cells:
        try:
            cell_validation = excel.validate_calcs(output_addrs=[cell])
            if not cell_validation:
                print(f"  ✅ {cell}: Validado")
            else:
                print(f"  ⚠️ {cell}: Problemas detectados")
        except Exception as e:
            print(f"  ❌ {cell}: Error - {e}")
    
    # Validación de serialización
    print("\n3. Validación de serialización...")
    try:
        # Crear sub-modelo para testing
        excel.trim_graph(
            input_addrs=['Assumptions!B1'],  # Revenue_Growth
            output_addrs=['Summary!B1']      # Revenue
        )
        
        serialization_results = excel.validate_serialized(
            output_addrs=['Summary!B1']
        )
        
        if not serialization_results:
            print("  ✅ Serialización consistente")
        else:
            print(f"  ⚠️ Problemas en serialización: {len(serialization_results)}")
            
    except Exception as e:
        print(f"  ❌ Error en validación de serialización: {e}")
    
    # Validación con tolerancia personalizada
    print("\n4. Validación con tolerancia personalizada...")
    try:
        tolerance_validation = excel.validate_calcs(
            output_addrs=['Summary!B1'],  # Revenue
            tolerance=0.01  # 1 centavo de tolerancia
        )
        
        if not tolerance_validation:
            print("  ✅ Validación con tolerancia: Aprobada")
        else:
            print("  ⚠️ Diferencias mayores a tolerancia detectadas")
            
    except Exception as e:
        print(f"  ❌ Error en validación con tolerancia: {e}")


def example_5_export_and_documentation():
    """
    Ejemplo 5: Exportación y Documentación
    
    Demuestra cómo exportar modelos en diferentes formatos
    y generar documentación automática.
    """
    print("\n=== Ejemplo 5: Exportación y Documentación ===")
    
    excel = create_sample_financial_model()
    
    # Crear sub-modelo para exportación
    excel.trim_graph(
        input_addrs=['Assumptions!B1', 'Assumptions!B2'],  # Revenue_Growth, COGS_Rate
        output_addrs=['Summary!B5']  # EBITDA
    )
    
    print(f"Modelo preparado para exportación: {len(excel.cell_map)} celdas")
    
    # Exportar en múltiples formatos
    with tempfile.TemporaryDirectory() as tmpdir:
        base_path = os.path.join(tmpdir, 'financial_model')
        
        print("\nExportando modelo en múltiples formatos:")
        
        # 1. Pickle (más rápido)
        try:
            pickle_file = f"{base_path}.pkl"
            excel.to_file(pickle_file)
            size_pkl = os.path.getsize(pickle_file)
            print(f"  ✅ Pickle: {pickle_file} ({size_pkl:,} bytes)")
        except Exception as e:
            print(f"  ❌ Error exportando Pickle: {e}")
        
        # 2. YAML (legible)
        try:
            yaml_file = f"{base_path}.yml"
            excel.to_file(yaml_file)
            size_yml = os.path.getsize(yaml_file)
            print(f"  ✅ YAML: {yaml_file} ({size_yml:,} bytes)")
        except Exception as e:
            print(f"  ❌ Error exportando YAML: {e}")
        
        # 3. JSON (portable)
        try:
            json_file = f"{base_path}.json"
            excel.to_file(json_file)
            size_json = os.path.getsize(json_file)
            print(f"  ✅ JSON: {json_file} ({size_json:,} bytes)")
        except Exception as e:
            print(f"  ❌ Error exportando JSON: {e}")
        
        # 4. GEXF para visualización
        try:
            gexf_file = f"{base_path}.gexf"
            excel.export_to_gexf(gexf_file)
            size_gexf = os.path.getsize(gexf_file)
            print(f"  ✅ GEXF: {gexf_file} ({size_gexf:,} bytes)")
        except Exception as e:
            print(f"  ❌ Error exportando GEXF: {e}")
    
    # Generar documentación de dependencias
    print("\nGenerando documentación de dependencias:")
    
    critical_outputs = ['Summary!B5']  # EBITDA
    
    for output in critical_outputs:
        print(f"\n--- Documentación para {output} ---")
        try:
            # Mostrar árbol de dependencias
            tree_lines = list(excel.value_tree_str(output))
            for line in tree_lines[:10]:  # Mostrar solo las primeras 10 líneas
                print(line)
            
            if len(tree_lines) > 10:
                print(f"... y {len(tree_lines) - 10} líneas más")
                
        except Exception as e:
            print(f"Error generando documentación: {e}")
    
    # Estadísticas del modelo
    print(f"\nEstadísticas del modelo:")
    print(f"  Total de celdas: {len(excel.cell_map)}")
    print(f"  Nodos en grafo: {len(excel.dep_graph.nodes())}")
    print(f"  Edges en grafo: {len(excel.dep_graph.edges())}")
    
    # Identificar celdas con más dependencias
    if excel.dep_graph.nodes():
        print(f"\nCeldas con más dependientes:")
        cell_deps = [(cell, len(list(excel.dep_graph.successors(cell)))) 
                     for cell in excel.dep_graph.nodes()]
        cell_deps.sort(key=lambda x: x[1], reverse=True)
        
        for cell, dep_count in cell_deps[:5]:
            print(f"  {cell.address.address}: {dep_count} dependientes")


def create_sample_financial_model():
    """
    Crea un modelo financiero de ejemplo para demostración.
    
    En un caso real, esto sería reemplazado por:
    return ExcelCompiler('path/to/financial_model.xlsx')
    """
    from openpyxl import Workbook
    
    wb = Workbook()
    
    # Hoja de Assumptions
    assumptions = wb.create_sheet('Assumptions')
    assumptions['A1'] = 'Revenue_Growth'
    assumptions['B1'] = 0.05  # 5%
    assumptions['A2'] = 'COGS_Rate'
    assumptions['B2'] = 0.60  # 60%
    assumptions['A3'] = 'OpEx_Rate'
    assumptions['B3'] = 0.25  # 25%
    assumptions['A4'] = 'Base_Revenue'
    assumptions['B4'] = 1000000  # $1M
    
    # Hoja de Summary
    summary = wb.create_sheet('Summary')
    summary['A1'] = 'Revenue'
    summary['B1'] = '=Assumptions!B4*(1+Assumptions!B1)'
    summary['A2'] = 'COGS'
    summary['B2'] = '=B1*Assumptions!B2'
    summary['A3'] = 'Gross_Profit'
    summary['B3'] = '=B1-B2'
    summary['A4'] = 'OpEx'
    summary['B4'] = '=B1*Assumptions!B3'
    summary['A5'] = 'EBITDA'
    summary['B5'] = '=B3-B4'
    summary['A6'] = 'FCF'
    summary['B6'] = '=B5*0.8'  # Simplified FCF
    
    # Remover hoja por defecto
    wb.remove(wb['Sheet'])
    
    return ExcelCompiler(excel=wb)


def main():
    """Ejecuta todos los ejemplos."""
    print("Ejemplos de Model Focusing en Pycel")
    print("=" * 50)
    
    try:
        example_1_financial_model_audit()
        example_2_sensitivity_analysis()
        example_3_dependency_analysis()
        example_4_model_validation()
        example_5_export_and_documentation()
        
        print("\n" + "=" * 50)
        print("✅ Todos los ejemplos ejecutados exitosamente")
        
    except Exception as e:
        print(f"\n❌ Error ejecutando ejemplos: {e}")
        import traceback
        traceback.print_exc()


if __name__ == '__main__':
    main()